"""localConnect — Flask backend + React SPA (single deployable service).

Serves the built React frontend (in ./frontend) at the root and exposes the
JSON API under /api/*. Run:
    pip install -r requirements.txt
    python app.py          # seeds the DB on first run, serves on :5000
Production:
    gunicorn app:app --bind 0.0.0.0:$PORT --workers 1 --timeout 120
"""

import os
from functools import wraps

from flask import (Flask, abort, jsonify, request, send_file,
                   send_from_directory)
from flask_cors import CORS
from flask_login import (LoginManager, current_user, login_required,
                         login_user, logout_user)

import analytics
import services
from config import Config
from models import (Inventory, Order, OrderItem, Payment, Product, Shop, User, db)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")  # built React SPA

app = Flask(__name__, static_folder=None)
app.config.from_object(Config)

# CORS only matters if the frontend is hosted separately; harmless when merged.
_origins = [o.strip() for o in os.environ.get(
    "CORS_ORIGINS", "http://localhost:5173,http://127.0.0.1:5173").split(",") if o.strip()]
CORS(app, resources={r"/api/*": {"origins": _origins}}, supports_credentials=True)
if os.environ.get("CORS_ORIGINS"):  # cross-site cookies need SameSite=None+Secure
    app.config.update(SESSION_COOKIE_SAMESITE="None", SESSION_COOKIE_SECURE=True)

db.init_app(app)

login_manager = LoginManager(app)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@login_manager.unauthorized_handler
def _unauthorized():
    return jsonify({"error": "Authentication required"}), 401


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def role_required(*roles):
    def decorator(fn):
        @wraps(fn)
        @login_required
        def wrapper(*args, **kwargs):
            if current_user.role not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def current_location():
    """Resolve the active customer location (query/cookie > profile > default)."""
    try:
        lat = float(request.args.get("lat") or request.cookies.get("lat"))
        lng = float(request.args.get("lng") or request.cookies.get("lng"))
        return lat, lng
    except (TypeError, ValueError):
        if current_user.is_authenticated and current_user.lat:
            return current_user.lat, current_user.lng
        return app.config["DEFAULT_LAT"], app.config["DEFAULT_LNG"]


def _require_my_shop():
    return Shop.query.filter_by(owner_id=current_user.id).first()


def _owner_shop_or_error():
    if not current_user.is_authenticated:
        return None, (jsonify({"error": "Not authenticated"}), 401)
    if current_user.role != "owner":
        return None, (jsonify({"error": "Shopkeeper account required"}), 403)
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return None, (jsonify({"error": "No shop"}), 404)
    return shop, None


def _my_order_or_error(order_id):
    if not current_user.is_authenticated:
        return None, (jsonify({"error": "Not authenticated"}), 401)
    order = db.session.get(Order, order_id)
    if not order or order.customer_id != current_user.id:
        return None, (jsonify({"error": "Order not found"}), 404)
    return order, None


def _user_json(user):
    shop = (Shop.query.filter_by(owner_id=user.id).first()
            if user.role == "owner" else None)
    return {
        "id": user.id, "name": user.name, "email": user.email,
        "phone": user.phone, "role": user.role,
        "shop_id": shop.id if shop else None,
        "shop_name": shop.name if shop else None,
    }


def _business_answer(shop_id, msg):
    m = msg.lower().strip()
    rev = analytics.revenue_analytics(shop_id)
    sales = analytics.sales_analytics(shop_id)
    perf = analytics.product_performance(shop_id)
    prof = analytics.profit_savings(shop_id)
    today = analytics.realtime_today(shop_id)
    pat = analytics.purchase_patterns(shop_id)

    if not m:
        return ("Ask me about revenue, profit, best sellers, low stock, today's "
                "orders, peak hours, or what to restock.")
    if any(w in m for w in ["today", "now", "live"]):
        return (f"📅 Today: {today['orders_today']} orders, ₹{today['revenue_today']:.0f} "
                f"revenue, {today['products_sold_today']} products sold, "
                f"{today['active_deliveries']} active deliveries.")
    if "profit" in m or "margin" in m:
        return (f"💰 Profit ₹{prof['profit']:.0f} on ₹{prof['revenue']:.0f} revenue "
                f"({prof['margin']}% margin). Inventory savings ₹{prof['inventory_savings']:.0f}.")
    if "revenue" in m or "sales" in m or "earn" in m:
        return (f"📈 Revenue — today ₹{rev['daily']:.0f}, this week ₹{rev['weekly']:.0f}, "
                f"this month ₹{rev['monthly']:.0f}, total ₹{rev['total']:.0f} "
                f"(growth {rev['growth']}%).")
    if any(w in m for w in ["best", "top", "popular", "selling"]):
        t = perf["top_selling"][:3]
        return "🏆 Top sellers: " + ", ".join(f"{r['name']} ({r['sold']})" for r in t) if t else "No sales yet."
    if any(w in m for w in ["low stock", "restock", "run out", "reorder"]):
        di = analytics.demand_insights(shop_id)
        if di["restock"]:
            return "🔁 Restock: " + ", ".join(f"{s['product']} (+{s['recommended_order']})"
                                              for s in di["restock"][:4])
        return "✅ Stock levels look healthy."
    if any(w in m for w in ["peak", "busy", "hour", "when", "day"]):
        return f"⏰ Peak hour {pat['peak_hour']}, busiest day {pat['peak_day']}."
    if any(w in m for w in ["order", "pending", "cancel"]):
        return (f"🧾 {sales['total_orders']} orders — {sales['completed']} completed, "
                f"{sales['pending']} pending, {sales['cancelled']} cancelled. "
                f"AOV ₹{sales['avg_order_value']:.0f}.")
    if "together" in m or "bundle" in m:
        fbt = pat["frequently_bought_together"][:3]
        return ("🧺 Frequently bought together: "
                + "; ".join(f"{a} + {b}" for a, b, _ in fbt)) if fbt else "Not enough data yet."
    return ("I can tell you about revenue, profit, top sellers, low stock/restock, "
            "today's activity, peak hours, orders, or product bundles.")


def _build_excel(shop):
    import io
    from openpyxl import Workbook
    d = analytics.shop_dashboard_data(shop.id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"{shop.name} — Analytics Report"]); ws.append([])
    ws.append(["Revenue"])
    for k, v in d["revenue"].items():
        ws.append([k.capitalize(), v])
    ws.append([]); ws.append(["Sales"])
    for k, v in d["sales"].items():
        ws.append([k.replace("_", " ").title(), v])
    ws.append([]); ws.append(["Profit & Savings"])
    for k, v in d["profit"].items():
        ws.append([k.replace("_", " ").title(), v])
    ws2 = wb.create_sheet("Top Products")
    ws2.append(["Product", "Category", "Units Sold", "Views", "Stock", "Price"])
    for r in d["performance"]["top_selling"]:
        ws2.append([r["name"], r["category"], r["sold"], r["views"], r["stock"], r["price"]])
    ws3 = wb.create_sheet("Restock Suggestions")
    ws3.append(["Product", "Current Stock", "Recommended Order"])
    for s in d["demand"]["restock"]:
        ws3.append([s["product"], s["current_stock"], s["recommended_order"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{shop.name.replace(' ', '_')}_analytics.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _build_pdf(shop):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)
    d = analytics.shop_dashboard_data(shop.id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    el = [Paragraph(f"{shop.name} — Business Analytics Report", styles["Title"]),
          Paragraph(f"Category: {shop.category} · Rating: {shop.rating}", styles["Normal"]),
          Spacer(1, 16)]

    def section(title, pairs):
        el.append(Paragraph(title, styles["Heading2"]))
        t = Table([[k, str(v)] for k, v in pairs], colWidths=[220, 220])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7e2")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2ff")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#f7f9fc")])]))
        el.append(t); el.append(Spacer(1, 14))

    r, s, p = d["revenue"], d["sales"], d["profit"]
    section("Revenue", [("Today", f"Rs {r['daily']:.0f}"), ("This Week", f"Rs {r['weekly']:.0f}"),
                        ("This Month", f"Rs {r['monthly']:.0f}"), ("This Year", f"Rs {r['yearly']:.0f}"),
                        ("Total", f"Rs {r['total']:.0f}"), ("Growth", f"{r['growth']}%")])
    section("Sales", [("Total Orders", s["total_orders"]), ("Completed", s["completed"]),
                      ("Pending", s["pending"]), ("Cancelled", s["cancelled"]),
                      ("Avg Order Value", f"Rs {s['avg_order_value']:.0f}")])
    section("Profit & Savings", [("Revenue", f"Rs {p['revenue']:.0f}"), ("Profit", f"Rs {p['profit']:.0f}"),
                                 ("Margin", f"{p['margin']}%"), ("Inventory Cost", f"Rs {p['inventory_cost']:.0f}"),
                                 ("Operational Savings", f"Rs {p['operational_savings']:.0f}"),
                                 ("Waste Reduction", f"Rs {p['waste_reduction']:.0f}")])
    el.append(Paragraph("Top Selling Products", styles["Heading2"]))
    rows = [["Product", "Sold", "Stock", "Price"]] + [
        [x["name"], x["sold"], x["stock"], f"Rs {x['price']:.0f}"]
        for x in d["performance"]["top_selling"][:8]]
    t = Table(rows, colWidths=[200, 80, 80, 100])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7e2")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9)]))
    el.append(t)
    doc.build(el); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{shop.name.replace(' ', '_')}_analytics.pdf",
                     mimetype="application/pdf")


# ===========================================================================
# JSON API
# ===========================================================================

@app.route("/api/search")
def api_search():
    query = request.args.get("q", "").strip()
    category = request.args.get("category", "All")
    sort_by = request.args.get("sort", "distance")
    lat, lng = current_location()
    results = services.search_inventory(query, lat, lng, category)
    results = (services.best_value_score(results) if sort_by == "best"
               else services.sort_results(results, sort_by))
    return jsonify({"count": len(results), "results": results})


@app.route("/api/shops")
def api_shops():
    lat, lng = current_location()
    shops = Shop.query.filter_by(verified=True).all()
    return jsonify([{
        "id": s.id, "name": s.name, "category": s.category,
        "lat": s.lat, "lng": s.lng, "rating": round(s.rating or 0, 1),
        "delivery": s.delivery_available, "address": s.address, "image": s.image,
        "distance_km": services.haversine_km(lat, lng, s.lat, s.lng),
        "maps_url": services.maps_link(s.name, s.address),
    } for s in shops])


@app.route("/api/suggest")
def api_suggest():
    lat, lng = current_location()
    return jsonify(services.suggest(request.args.get("q", ""), lat, lng))


@app.route("/api/home")
def api_home():
    return jsonify({
        "popular": services.popular_products(limit=10),
        "categories": ["Clothing", "Hardware", "Electronics", "Grocery",
                       "Footwear", "Stationery", "Sports", "Home Decor"],
        "stats": {"shops": Shop.query.count(), "products": Product.query.count(),
                  "orders": Order.query.count()},
    })


@app.route("/api/shop/<int:shop_id>")
def api_shop(shop_id):
    s = db.session.get(Shop, shop_id) or abort(404)
    lat, lng = current_location()
    return jsonify({
        "id": s.id, "name": s.name, "category": s.category, "address": s.address,
        "contact": s.mobile, "rating": round(s.rating or 0, 1),
        "delivery": s.delivery_available, "parking": s.parking, "image": s.image,
        "lat": s.lat, "lng": s.lng, "verified": s.verified,
        "distance_km": services.haversine_km(lat, lng, s.lat, s.lng),
        "maps_url": services.maps_link(s.name, s.address),
        "products": [{"inventory_id": i.id, "name": i.product.name,
                      "category": i.product.category, "image": i.product.image,
                      "price": i.price, "stock": i.stock, "status": i.status}
                     for i in s.inventory],
    })


@app.route("/api/shop/assistant", methods=["POST"])
@role_required("owner")
def shop_assistant():
    shop = _require_my_shop()
    if not shop:
        return jsonify({"reply": "Register your shop first."})
    return jsonify({"reply": _business_answer(shop.id, (request.json or {}).get("message", ""))})


@app.route("/api/shop/<int:shop_id>/forecast")
@role_required("owner", "admin")
def api_forecast(shop_id):
    return jsonify({"forecast": services.demand_forecast(shop_id),
                    "restock": services.restock_suggestions(shop_id)})


# ---- auth ----

@app.route("/api/auth/me")
def api_me():
    return jsonify(_user_json(current_user) if current_user.is_authenticated else None)


@app.route("/api/auth/register", methods=["POST"])
def api_register():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email and password are required."}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "That email is already registered."}), 409
    role = data.get("role", "customer")
    if role not in ("customer", "owner"):
        role = "customer"
    user = User(name=(data.get("name") or email.split("@")[0]).strip(),
                email=email, phone=(data.get("phone") or "").strip(), role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    login_user(user)
    return jsonify(_user_json(user))


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or {}
    user = User.query.filter_by(email=(data.get("email") or "").strip().lower()).first()
    if user and user.check_password(data.get("password") or ""):
        login_user(user)
        return jsonify(_user_json(user))
    return jsonify({"error": "Invalid email or password."}), 401


@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    logout_user()
    return jsonify({"ok": True})


# ---- orders ----

@app.route("/api/order", methods=["POST"])
def api_create_order():
    if not current_user.is_authenticated:
        return jsonify({"error": "Please log in to order."}), 401
    if current_user.role != "customer":
        return jsonify({"error": "Log in with a customer account to order."}), 403
    data = request.get_json(silent=True) or {}
    inv = db.session.get(Inventory, data.get("inventory_id"))
    if not inv:
        return jsonify({"error": "Product not found."}), 404
    qty = max(1, int(data.get("qty", 1)))
    if qty > inv.stock:
        return jsonify({"error": f"Only {inv.stock} in stock."}), 400
    mode = data.get("mode", "pickup")
    note = (data.get("note") or "").strip()
    addr = (data.get("address") or "").strip()
    payment = data.get("payment", "cod")
    if mode == "delivery":
        order_type, where = "delivery", (addr or "Home delivery")
    else:
        order_type = "reserve"
        where = "In-store pickup" + (f" — {note}" if note else "")
    order = Order(customer_id=current_user.id, shop_id=inv.shop_id,
                  order_type=order_type, status="placed",
                  total=round(inv.price * qty, 2), address=where)
    db.session.add(order)
    db.session.flush()
    db.session.add(OrderItem(order_id=order.id, inventory_id=inv.id,
                             product_name=inv.product.name, qty=qty, price=inv.price))
    inv.stock = max(0, inv.stock - qty)
    inv.sold = (inv.sold or 0) + qty
    db.session.add(Payment(order_id=order.id, method=payment, amount=order.total,
                           status="paid" if payment != "cod" else "pending",
                           txn_ref=f"TXN{order.id:06d}"))
    if order_type == "delivery":
        partner = services.assign_delivery_partner(inv.shop)
        if partner:
            order.delivery_partner_id = partner.id
            order.status = "accepted"
    db.session.commit()
    return jsonify({
        "id": order.id, "status": order.status, "type": order.order_type,
        "total": order.total, "qty": qty, "product": inv.product.name,
        "shop": inv.shop.name, "where": where,
        "message": ("Order placed for home delivery!" if order_type == "delivery"
                    else "Reserved! The shop has been notified to keep it ready for pickup."),
    })


@app.route("/api/order/<int:order_id>/action", methods=["POST"])
def api_order_action(order_id):
    shop, err = _owner_shop_or_error()
    if err:
        return err
    order = db.session.get(Order, order_id)
    if not order or order.shop_id != shop.id:
        return jsonify({"error": "Order not found"}), 404
    action = (request.get_json(silent=True) or {}).get("action", "")
    transitions = {
        "accept": "accepted", "reject": "rejected", "pack": "packed",
        "dispatch": "out_for_delivery", "deliver": "delivered", "collect": "collected",
        "placed": "placed", "accepted": "accepted", "packed": "packed",
        "out_for_delivery": "out_for_delivery", "delivered": "delivered",
    }
    if action not in transitions:
        return jsonify({"error": "Invalid action"}), 400
    order.status = transitions[action]
    db.session.commit()
    return jsonify({"id": order.id, "status": order.status})


@app.route("/api/my/order/<int:order_id>/cancel", methods=["POST"])
def api_cancel_order(order_id):
    order, err = _my_order_or_error(order_id)
    if err:
        return err
    if order.status in ("delivered", "collected", "cancelled", "rejected"):
        return jsonify({"error": f"This order can't be cancelled ({order.status})."}), 400
    for it in order.items:
        inv = db.session.get(Inventory, it.inventory_id) if it.inventory_id else None
        if inv:
            inv.stock = (inv.stock or 0) + it.qty
            inv.sold = max(0, (inv.sold or 0) - it.qty)
    order.status = "cancelled"
    if order.payment:
        order.payment.status = "cancelled"
    db.session.commit()
    return jsonify({"id": order.id, "status": "cancelled"})


@app.route("/api/my/order/<int:order_id>/remove", methods=["POST", "DELETE"])
def api_remove_order(order_id):
    order, err = _my_order_or_error(order_id)
    if err:
        return err
    if order.status not in ("delivered", "collected", "cancelled", "rejected"):
        for it in order.items:
            inv = db.session.get(Inventory, it.inventory_id) if it.inventory_id else None
            if inv:
                inv.stock = (inv.stock or 0) + it.qty
                inv.sold = max(0, (inv.sold or 0) - it.qty)
    db.session.delete(order)
    db.session.commit()
    return jsonify({"ok": True, "id": order_id})


@app.route("/api/my/orders")
def api_my_orders_json():
    if not current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    orders = sorted(current_user.orders,
                    key=lambda o: o.created_at.timestamp() if o.created_at else 0,
                    reverse=True)
    return jsonify([{
        "id": o.id, "status": o.status, "type": o.order_type, "total": round(o.total, 2),
        "shop": o.shop.name if o.shop else "—",
        "shop_maps_url": services.maps_link(o.shop.name, o.shop.address) if o.shop else None,
        "date": (o.created_at.strftime("%d %b %Y, %H:%M") if o.created_at else ""),
        "qty": sum(it.qty for it in o.items),
        "payment": (o.payment.method.upper() if o.payment else None),
        "items": [{"name": it.product_name, "qty": it.qty, "price": round(it.price, 2)}
                  for it in o.items],
    } for o in orders])


@app.route("/api/my/shop")
def api_my_shop():
    if not current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    if current_user.role != "owner":
        return jsonify({"error": "Shopkeeper account required"}), 403
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return jsonify({"shop": None})
    return jsonify({
        "shop": {"id": shop.id, "name": shop.name, "category": shop.category,
                 "rating": shop.rating, "address": shop.address,
                 "verified": shop.verified, "delivery": shop.delivery_available},
        "revenue": analytics.revenue_analytics(shop.id),
        "sales": analytics.sales_analytics(shop.id),
        "today": analytics.realtime_today(shop.id),
        "top_selling": analytics.product_performance(shop.id)["top_selling"][:6],
        "low_stock": analytics.product_performance(shop.id)["low_stock"][:8],
        "products": [{"name": i.product.name, "price": i.price, "stock": i.stock,
                      "status": i.status, "image": i.product.image}
                     for i in shop.inventory],
    })


@app.route("/api/my/analytics")
def api_my_analytics():
    shop, err = _owner_shop_or_error()
    if err:
        return err
    data = analytics.shop_dashboard_data(shop.id)
    data["shop"] = {"id": shop.id, "name": shop.name, "category": shop.category,
                    "rating": shop.rating, "address": shop.address,
                    "contact": shop.mobile, "lat": shop.lat, "lng": shop.lng,
                    "maps_url": services.maps_link(shop.name, shop.address),
                    "verified": shop.verified, "delivery": shop.delivery_available}
    orders = sorted(shop.orders,
                    key=lambda o: o.created_at.timestamp() if o.created_at else 0,
                    reverse=True)

    def _odist(o):
        if o.customer and o.customer.lat:
            return services.humanize_distance(
                services.haversine_km(shop.lat, shop.lng, o.customer.lat, o.customer.lng))
        return "—"

    data["orders"] = [{
        "id": o.id, "status": o.status, "type": o.order_type, "total": round(o.total, 2),
        "date": (o.created_at.strftime("%d %b %Y, %H:%M") if o.created_at else ""),
        "qty": sum(it.qty for it in o.items),
        "items": [{"name": it.product_name, "qty": it.qty} for it in o.items],
        "customer": (o.customer.name if o.customer else "Guest"),
        "distance": _odist(o),
        "payment": (o.payment.method.upper() if o.payment else "—"),
    } for o in orders[:40]]
    data["order_counts"] = analytics.sales_analytics(shop.id)

    in_stock, low_stock, out_stock = [], [], []
    for i in shop.inventory:
        row = {"name": i.product.name, "stock": i.stock, "price": i.price,
               "image": i.product.image}
        (out_stock if i.stock <= 0 else low_stock if i.stock <= Inventory.LOW_STOCK_THRESHOLD
         else in_stock).append(row)
    data["stock"] = {
        "in_stock": sorted(in_stock, key=lambda r: -r["stock"]),
        "low_stock": sorted(low_stock, key=lambda r: r["stock"]),
        "out_of_stock": out_stock,
        "counts": {"in_stock": len(in_stock), "low_stock": len(low_stock),
                   "out_of_stock": len(out_stock),
                   "total_units": sum(i.stock for i in shop.inventory)},
    }
    return jsonify(data)


@app.route("/api/my/export.xlsx")
def api_my_export_excel():
    shop, err = _owner_shop_or_error()
    return err if err else _build_excel(shop)


@app.route("/api/my/export.pdf")
def api_my_export_pdf():
    shop, err = _owner_shop_or_error()
    return err if err else _build_pdf(shop)


@app.route("/api/recommendations")
@login_required
def api_recommendations():
    return jsonify(services.recommend_for_user(current_user.id))


# ===========================================================================
# Serve the built React SPA (everything that isn't /api/*)
# ===========================================================================

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def spa(path):
    if path.startswith("api/") or path.startswith("api"):
        abort(404)
    candidate = os.path.join(FRONTEND_DIR, path)
    if path and os.path.isfile(candidate):
        return send_from_directory(FRONTEND_DIR, path)
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Not found"}), 404
    return send_from_directory(FRONTEND_DIR, "index.html")  # SPA client routes


@app.errorhandler(403)
def forbidden(e):
    return jsonify({"error": "Forbidden"}), 403


# ---------------------------------------------------------------------------

def init_db():
    """Create tables and seed sample data on first run (race-safe for gunicorn)."""
    with app.app_context():
        try:
            db.create_all()
            if Shop.query.count() == 0:
                import seed
                seed.run(db)
                print("✓ Database seeded with sample stores, products & users.")
        except Exception as e:
            db.session.rollback()
            print(f"⚠ init_db skipped (already initialising?): {e}")


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    print(f"\n🛒  localConnect running on 0.0.0.0:{port}")
    print("    Demo logins → owner1@localfind.in / customer@localfind.in   (password: password123)\n")
    app.run(host="0.0.0.0", port=port, debug=debug)
