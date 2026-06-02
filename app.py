"""Hyperlocal Commerce Platform — Flask application entry point.

A hyperlocal product-discovery marketplace: customers search for products and
instantly see nearby shops stocking them, with price, live stock, distance,
walking time, delivery options and ratings — plus comparison, maps, ordering,
payment, tracking, shop dashboards, an admin console and AI features.

Run:
    pip install -r requirements.txt
    python app.py            # seeds the DB on first run, serves on :5000
"""

from functools import wraps

from flask import (Flask, abort, flash, jsonify, redirect, render_template,
                   request, send_file, url_for)
from flask_login import (LoginManager, current_user, login_required,
                         login_user, logout_user)

import analytics
import services
from config import Config
from models import (DeliveryPartner, Inventory, Order, OrderItem, Payment,
                    Product, Review, Shop, User, db)

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message_category = "warning"


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.context_processor
def inject_globals():
    return {
        "maps_key": app.config["GOOGLE_MAPS_API_KEY"],
        "categories": ["Clothing", "Hardware", "Electronics", "Grocery",
                       "Footwear", "Stationery", "Sports", "Home Decor"],
    }


# ---------------------------------------------------------------------------
# Role guard
# ---------------------------------------------------------------------------

def role_required(*roles):
    def decorator(fn):
        @wraps(fn)
        @login_required
        def wrapper(*args, **kwargs):
            if current_user.role not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def current_location():
    """Resolve the active customer location (cookie/query > profile > default)."""
    try:
        lat = float(request.args.get("lat") or request.cookies.get("lat"))
        lng = float(request.args.get("lng") or request.cookies.get("lng"))
        return lat, lng
    except (TypeError, ValueError):
        if current_user.is_authenticated and current_user.lat:
            return current_user.lat, current_user.lng
        return app.config["DEFAULT_LAT"], app.config["DEFAULT_LNG"]


# ===========================================================================
# PUBLIC PAGES + AUTH
# ===========================================================================

@app.route("/")
def index():
    shops = Shop.query.filter_by(verified=True).limit(8).all()
    popular = services.popular_products(limit=8)
    return render_template("index.html", shops=shops, popular=popular,
                           stats={
                               "shops": Shop.query.count(),
                               "products": Product.query.count(),
                               "orders": Order.query.count(),
                           })


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        if User.query.filter_by(email=email).first():
            flash("That email is already registered.", "danger")
            return redirect(url_for("register"))
        user = User(
            name=request.form["name"].strip(),
            email=email,
            phone=request.form.get("phone", "").strip(),
            role=request.form.get("role", "customer"),
        )
        user.set_password(request.form["password"])
        db.session.add(user)
        db.session.commit()
        login_user(user)
        flash("Welcome to localConnect! Your account is ready.", "success")
        return redirect(url_for("post_login_redirect"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(request.form["password"]):
            login_user(user)
            return redirect(url_for("post_login_redirect"))
        flash("Invalid email or password.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))


@app.route("/go")
@login_required
def post_login_redirect():
    if current_user.role == "admin":
        return redirect(url_for("admin_dashboard"))
    if current_user.role == "owner":
        return redirect(url_for("shop_dashboard"))
    return redirect(url_for("customer_dashboard"))


# ===========================================================================
# CUSTOMER MODULE
# ===========================================================================

@app.route("/dashboard")
@role_required("customer")
def customer_dashboard():
    lat, lng = current_location()
    recs = services.recommend_for_user(current_user.id)
    popular = services.popular_products(limit=8)
    # nearby shops decorated with distance
    shops = []
    for s in Shop.query.filter_by(verified=True).all():
        d = services.haversine_km(lat, lng, s.lat, s.lng)
        shops.append((d, s))
    shops.sort(key=lambda x: x[0] or 1e9)
    nearby = [s for _, s in shops[:9]]
    orders = (Order.query.filter_by(customer_id=current_user.id)
              .order_by(Order.created_at.desc()).limit(5).all())
    return render_template("customer/dashboard.html", recs=recs, popular=popular,
                           nearby=nearby, orders=orders, lat=lat, lng=lng)


@app.route("/search")
def search():
    """Search Results Page (works for guests too — encourages signup)."""
    query = request.args.get("q", "").strip()
    category = request.args.get("category", "All")
    sort_by = request.args.get("sort", "distance")
    lat, lng = current_location()

    results = services.search_inventory(query, lat, lng, category)
    if sort_by == "best":
        results = services.best_value_score(results)
    else:
        results = services.sort_results(results, sort_by)

    return render_template("customer/search.html", results=results, query=query,
                           category=category, sort_by=sort_by, lat=lat, lng=lng)


@app.route("/compare")
def compare():
    """Product Comparison Page — same product across shops, side by side."""
    query = request.args.get("q", "").strip()
    lat, lng = current_location()
    results = services.best_value_score(
        services.search_inventory(query, lat, lng))
    return render_template("customer/compare.html", results=results,
                           query=query, lat=lat, lng=lng)


@app.route("/map")
def map_view():
    """Google Maps Page — all shops + user, with routing/navigation."""
    lat, lng = current_location()
    shops = Shop.query.filter_by(verified=True).all()
    shop_data = [{
        "id": s.id, "name": s.name, "category": s.category,
        "lat": s.lat, "lng": s.lng, "rating": round(s.rating or 0, 1),
        "address": s.address, "delivery": s.delivery_available,
        "distance": services.humanize_distance(
            services.haversine_km(lat, lng, s.lat, s.lng)),
        "maps_url": services.maps_link(s.name, s.address),
    } for s in shops]
    return render_template("customer/map.html", shops=shop_data, lat=lat, lng=lng)


@app.route("/checkout/<int:inventory_id>", methods=["GET", "POST"])
@role_required("customer")
def checkout(inventory_id):
    inv = db.session.get(Inventory, inventory_id) or abort(404)
    lat, lng = current_location()

    if request.method == "POST":
        qty = max(1, int(request.form.get("qty", 1)))
        order_type = request.form.get("order_type", "reserve")
        method = request.form.get("payment_method", "cod")

        if qty > inv.stock:
            flash("Not enough stock available.", "danger")
            return redirect(url_for("checkout", inventory_id=inventory_id))

        order = Order(
            customer_id=current_user.id,
            shop_id=inv.shop_id,
            order_type=order_type,
            status="placed",
            total=round(inv.price * qty, 2),
            address=request.form.get("address", ""),
        )
        db.session.add(order)
        db.session.flush()

        db.session.add(OrderItem(order_id=order.id, inventory_id=inv.id,
                                 product_name=inv.product.name,
                                 qty=qty, price=inv.price))

        # decrement stock / record sale (drives AI popularity)
        inv.stock = max(0, inv.stock - qty)
        inv.sold = (inv.sold or 0) + qty

        # payment
        payment = Payment(order_id=order.id, method=method, amount=order.total,
                          status="paid" if method != "cod" else "pending",
                          txn_ref=f"TXN{order.id:06d}")
        db.session.add(payment)

        # delivery assignment
        if order_type == "delivery":
            partner = services.assign_delivery_partner(inv.shop)
            if partner:
                order.delivery_partner_id = partner.id
                order.status = "accepted"

        db.session.commit()
        flash("Order placed successfully!", "success")
        return redirect(url_for("track_order", order_id=order.id))

    dist = services.haversine_km(lat, lng, inv.shop.lat, inv.shop.lng)
    return render_template("customer/checkout.html", inv=inv,
                           distance=services.humanize_distance(dist),
                           delivery_time=services.delivery_time_min(dist),
                           walk_time=services.walking_time_min(dist))


@app.route("/orders")
@role_required("customer")
def my_orders():
    orders = (Order.query.filter_by(customer_id=current_user.id)
              .order_by(Order.created_at.desc()).all())
    return render_template("customer/orders.html", orders=orders)


@app.route("/track/<int:order_id>")
@role_required("customer", "admin")
def track_order(order_id):
    order = db.session.get(Order, order_id) or abort(404)
    if current_user.role == "customer" and order.customer_id != current_user.id:
        abort(403)
    # timeline stages for the tracking UI
    flow = (["placed", "accepted", "out_for_delivery", "delivered"]
            if order.order_type == "delivery"
            else ["placed", "accepted", "collected"])
    return render_template("customer/track.html", order=order, flow=flow)


@app.route("/shop/<int:shop_id>/review", methods=["POST"])
@role_required("customer")
def add_review(shop_id):
    shop = db.session.get(Shop, shop_id) or abort(404)
    review = Review(shop_id=shop.id, customer_id=current_user.id,
                    rating=int(request.form.get("rating", 5)),
                    comment=request.form.get("comment", "").strip())
    db.session.add(review)
    db.session.flush()
    # recompute shop average rating
    ratings = [r.rating for r in shop.reviews]
    shop.rating = round(sum(ratings) / len(ratings), 2) if ratings else 4.0
    db.session.commit()
    flash("Thanks for your review!", "success")
    return redirect(request.referrer or url_for("customer_dashboard"))


# ===========================================================================
# SHOP OWNER MODULE
# ===========================================================================

@app.route("/shop")
@role_required("owner")
def shop_dashboard():
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return redirect(url_for("shop_register"))

    orders = (Order.query.filter_by(shop_id=shop.id)
              .order_by(Order.created_at.desc()).all())
    revenue = sum(o.total for o in orders if o.status in
                  ("delivered", "collected", "accepted", "out_for_delivery"))
    low_stock = [i for i in shop.inventory if i.status == "low_stock"]
    out_stock = [i for i in shop.inventory if i.status == "out_of_stock"]

    analytics = {
        "total_orders": len(orders),
        "revenue": round(revenue, 2),
        "popular": services.popular_products(limit=5, shop_id=shop.id),
        "reviews": (Review.query.filter_by(shop_id=shop.id)
                    .order_by(Review.created_at.desc()).limit(5).all()),
        "forecast": services.demand_forecast(shop.id)[:5],
        "restock": services.restock_suggestions(shop.id),
    }
    return render_template("shop/dashboard.html", shop=shop, orders=orders,
                           low_stock=low_stock, out_stock=out_stock,
                           analytics=analytics)


@app.route("/shop/register", methods=["GET", "POST"])
@role_required("owner")
def shop_register():
    if Shop.query.filter_by(owner_id=current_user.id).first():
        return redirect(url_for("shop_dashboard"))
    if request.method == "POST":
        shop = Shop(
            owner_id=current_user.id,
            name=request.form["name"].strip(),
            owner_name=request.form.get("owner_name", current_user.name),
            mobile=request.form.get("mobile", ""),
            address=request.form.get("address", ""),
            lat=float(request.form.get("lat") or app.config["DEFAULT_LAT"]),
            lng=float(request.form.get("lng") or app.config["DEFAULT_LNG"]),
            category=request.form.get("category", "Grocery"),
            delivery_available=request.form.get("delivery_available") == "on",
            delivery_radius=float(request.form.get("delivery_radius", 3)),
            parking=request.form.get("parking") == "on",
            verified=False,
        )
        db.session.add(shop)
        db.session.commit()
        flash("Shop submitted! It will appear once an admin verifies it.",
              "success")
        return redirect(url_for("shop_dashboard"))
    return render_template("shop/register.html")


@app.route("/shop/product/add", methods=["POST"])
@role_required("owner")
def add_product():
    shop = Shop.query.filter_by(owner_id=current_user.id).first() or abort(404)
    name = request.form["name"].strip()
    category = request.form.get("category", shop.category)
    product = (Product.query.filter(Product.name.ilike(name)).first()
               or Product(name=name, category=category,
                          image=request.form.get("image", "")))
    if not product.id:
        db.session.add(product)
        db.session.flush()
    db.session.add(Inventory(shop_id=shop.id, product_id=product.id,
                             price=float(request.form["price"]),
                             stock=int(request.form.get("stock", 0))))
    db.session.commit()
    flash(f"Added {name} to your inventory.", "success")
    return redirect(url_for("shop_dashboard"))


@app.route("/shop/inventory/<int:inv_id>/update", methods=["POST"])
@role_required("owner")
def update_stock(inv_id):
    inv = db.session.get(Inventory, inv_id) or abort(404)
    if inv.shop.owner_id != current_user.id:
        abort(403)
    inv.stock = int(request.form.get("stock", inv.stock))
    if request.form.get("price"):
        inv.price = float(request.form["price"])
    db.session.commit()
    flash("Inventory updated.", "success")
    return redirect(url_for("shop_dashboard"))


@app.route("/shop/order/<int:order_id>/<action>", methods=["POST"])
@role_required("owner")
def order_action(order_id, action):
    order = db.session.get(Order, order_id) or abort(404)
    if order.shop.owner_id != current_user.id:
        abort(403)
    transitions = {
        "accept": "accepted", "reject": "rejected",
        "dispatch": "out_for_delivery", "deliver": "delivered",
        "collect": "collected",
    }
    if action in transitions:
        order.status = transitions[action]
        db.session.commit()
        flash(f"Order #{order.id} marked {order.status}.", "success")
    return redirect(url_for("shop_dashboard"))


# ---------------------------------------------------------------------------
# SHOP OWNER · BUSINESS-INTELLIGENCE ANALYTICS DASHBOARD  (owner-only)
# ---------------------------------------------------------------------------

def _require_my_shop():
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return None
    return shop


@app.route("/shop/analytics")
@role_required("owner")
def shop_analytics():
    shop = _require_my_shop()
    if not shop:
        return redirect(url_for("shop_register"))
    data = analytics.shop_dashboard_data(shop.id)
    return render_template("shop/analytics.html", shop=shop, d=data)


@app.route("/api/shop/assistant", methods=["POST"])
@role_required("owner")
def shop_assistant():
    """Conversational business assistant over the owner's own analytics."""
    shop = _require_my_shop()
    if not shop:
        return jsonify({"reply": "Register your shop first."})
    msg = (request.json or {}).get("message", "")
    return jsonify({"reply": _business_answer(shop.id, msg)})


def _business_answer(shop_id, msg):
    m = msg.lower().strip()
    rev = analytics.revenue_analytics(shop_id)
    sales = analytics.sales_analytics(shop_id)
    perf = analytics.product_performance(shop_id)
    prof = analytics.profit_savings(shop_id)
    today = analytics.realtime_today(shop_id)
    pat = analytics.purchase_patterns(shop_id)

    if not m:
        return ("Ask me about revenue, profit, best sellers, low stock, today's "
                "orders, peak hours, or what to restock.")
    if any(w in m for w in ["today", "now", "live"]):
        return (f"📅 Today: {today['orders_today']} orders, ₹{today['revenue_today']:.0f} "
                f"revenue, {today['products_sold_today']} products sold, "
                f"{today['active_deliveries']} active deliveries.")
    if "profit" in m or "margin" in m:
        return (f"💰 Profit ₹{prof['profit']:.0f} on ₹{prof['revenue']:.0f} revenue "
                f"({prof['margin']}% margin). Inventory savings ₹{prof['inventory_savings']:.0f}.")
    if "revenue" in m or "sales" in m or "earn" in m:
        return (f"📈 Revenue — today ₹{rev['daily']:.0f}, this week ₹{rev['weekly']:.0f}, "
                f"this month ₹{rev['monthly']:.0f}, total ₹{rev['total']:.0f} "
                f"(growth {rev['growth']}%).")
    if any(w in m for w in ["best", "top", "popular", "selling"]):
        t = perf["top_selling"][:3]
        return "🏆 Top sellers: " + ", ".join(f"{r['name']} ({r['sold']})" for r in t) if t else "No sales yet."
    if any(w in m for w in ["low stock", "restock", "run out", "reorder"]):
        di = analytics.demand_insights(shop_id)
        if di["restock"]:
            return "🔁 Restock: " + ", ".join(f"{s['product']} (+{s['recommended_order']})"
                                              for s in di["restock"][:4])
        return "✅ Stock levels look healthy."
    if any(w in m for w in ["peak", "busy", "hour", "when", "day"]):
        return f"⏰ Peak hour {pat['peak_hour']}, busiest day {pat['peak_day']}."
    if any(w in m for w in ["order", "pending", "cancel"]):
        return (f"🧾 {sales['total_orders']} orders — {sales['completed']} completed, "
                f"{sales['pending']} pending, {sales['cancelled']} cancelled. "
                f"AOV ₹{sales['avg_order_value']:.0f}.")
    if "together" in m or "bundle" in m:
        fbt = pat["frequently_bought_together"][:3]
        return ("🧺 Frequently bought together: "
                + "; ".join(f"{a} + {b}" for a, b, _ in fbt)) if fbt else "Not enough data yet."
    return ("I can tell you about revenue, profit, top sellers, low stock/restock, "
            "today's activity, peak hours, orders, or product bundles.")


@app.route("/shop/analytics/export.xlsx")
@role_required("owner")
def export_excel():
    shop = _require_my_shop() or abort(404)
    return _build_excel(shop)


@app.route("/shop/analytics/export.pdf")
@role_required("owner")
def export_pdf():
    shop = _require_my_shop() or abort(404)
    return _build_pdf(shop)


def _build_excel(shop):
    import io
    from openpyxl import Workbook

    d = analytics.shop_dashboard_data(shop.id)
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append([f"{shop.name} — Analytics Report"])
    ws.append([])
    ws.append(["Revenue"])
    for k, v in d["revenue"].items():
        ws.append([k.capitalize(), v])
    ws.append([])
    ws.append(["Sales"])
    for k, v in d["sales"].items():
        ws.append([k.replace("_", " ").title(), v])
    ws.append([])
    ws.append(["Profit & Savings"])
    for k, v in d["profit"].items():
        ws.append([k.replace("_", " ").title(), v])

    ws2 = wb.create_sheet("Top Products")
    ws2.append(["Product", "Category", "Units Sold", "Views", "Stock", "Price"])
    for r in d["performance"]["top_selling"]:
        ws2.append([r["name"], r["category"], r["sold"], r["views"], r["stock"], r["price"]])

    ws3 = wb.create_sheet("Restock Suggestions")
    ws3.append(["Product", "Current Stock", "Recommended Order"])
    for s in d["demand"]["restock"]:
        ws3.append([s["product"], s["current_stock"], s["recommended_order"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{shop.name.replace(' ', '_')}_analytics.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _build_pdf(shop):
    import io

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    d = analytics.shop_dashboard_data(shop.id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    el = [Paragraph(f"{shop.name} — Business Analytics Report", styles["Title"]),
          Paragraph(f"Category: {shop.category} · Rating: {shop.rating}", styles["Normal"]),
          Spacer(1, 16)]

    def section(title, pairs):
        el.append(Paragraph(title, styles["Heading2"]))
        t = Table([[k, str(v)] for k, v in pairs], colWidths=[220, 220])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7e2")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef2ff")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (1, 0), (1, -1), [colors.white, colors.HexColor("#f7f9fc")]),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

    r, s, p = d["revenue"], d["sales"], d["profit"]
    section("Revenue", [("Today", f"Rs {r['daily']:.0f}"), ("This Week", f"Rs {r['weekly']:.0f}"),
                        ("This Month", f"Rs {r['monthly']:.0f}"), ("This Year", f"Rs {r['yearly']:.0f}"),
                        ("Total", f"Rs {r['total']:.0f}"), ("Growth", f"{r['growth']}%")])
    section("Sales", [("Total Orders", s["total_orders"]), ("Completed", s["completed"]),
                      ("Pending", s["pending"]), ("Cancelled", s["cancelled"]),
                      ("Avg Order Value", f"Rs {s['avg_order_value']:.0f}")])
    section("Profit & Savings", [("Revenue", f"Rs {p['revenue']:.0f}"), ("Profit", f"Rs {p['profit']:.0f}"),
                                 ("Margin", f"{p['margin']}%"), ("Inventory Cost", f"Rs {p['inventory_cost']:.0f}"),
                                 ("Operational Savings", f"Rs {p['operational_savings']:.0f}"),
                                 ("Waste Reduction", f"Rs {p['waste_reduction']:.0f}")])

    el.append(Paragraph("Top Selling Products", styles["Heading2"]))
    rows = [["Product", "Sold", "Stock", "Price"]] + [
        [x["name"], x["sold"], x["stock"], f"Rs {x['price']:.0f}"]
        for x in d["performance"]["top_selling"][:8]]
    t = Table(rows, colWidths=[200, 80, 80, 100])
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7e2")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9)]))
    el.append(t)

    doc.build(el)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{shop.name.replace(' ', '_')}_analytics.pdf",
                     mimetype="application/pdf")


# ===========================================================================
# ADMIN MODULE
# ===========================================================================

@app.route("/admin")
@role_required("admin")
def admin_dashboard():
    return render_template(
        "admin/dashboard.html",
        users=User.query.all(),
        shops=Shop.query.all(),
        products=Product.query.all(),
        orders=Order.query.order_by(Order.created_at.desc()).all(),
        stats={
            "users": User.query.count(),
            "shops": Shop.query.count(),
            "pending_shops": Shop.query.filter_by(verified=False).count(),
            "orders": Order.query.count(),
            "revenue": round(sum(o.total for o in Order.query.all()), 2),
        },
    )


@app.route("/admin/analytics")
@role_required("admin")
def admin_analytics():
    return render_template("admin/analytics.html", d=analytics.admin_dashboard_data())


@app.route("/admin/shop/<int:shop_id>/verify", methods=["POST"])
@role_required("admin")
def verify_shop(shop_id):
    shop = db.session.get(Shop, shop_id) or abort(404)
    shop.verified = not shop.verified
    db.session.commit()
    flash(f"{shop.name} {'verified' if shop.verified else 'unverified'}.",
          "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/delete", methods=["POST"])
@role_required("admin")
def delete_user(user_id):
    user = db.session.get(User, user_id) or abort(404)
    if user.role == "admin":
        flash("Cannot delete an admin account.", "danger")
    else:
        db.session.delete(user)
        db.session.commit()
        flash("User removed.", "success")
    return redirect(url_for("admin_dashboard"))


# ===========================================================================
# JSON API ENDPOINTS (consumed by the frontend JS / maps / search-as-you-type)
# ===========================================================================

@app.route("/api/search")
def api_search():
    query = request.args.get("q", "").strip()
    category = request.args.get("category", "All")
    sort_by = request.args.get("sort", "distance")
    lat, lng = current_location()
    results = services.search_inventory(query, lat, lng, category)
    results = (services.best_value_score(results) if sort_by == "best"
               else services.sort_results(results, sort_by))
    return jsonify({"count": len(results), "results": results})


@app.route("/api/shops")
def api_shops():
    lat, lng = current_location()
    shops = Shop.query.filter_by(verified=True).all()
    return jsonify([{
        "id": s.id, "name": s.name, "category": s.category,
        "lat": s.lat, "lng": s.lng, "rating": round(s.rating or 0, 1),
        "delivery": s.delivery_available, "address": s.address,
        "image": s.image,
        "distance_km": services.haversine_km(lat, lng, s.lat, s.lng),
        "maps_url": services.maps_link(s.name, s.address),
    } for s in shops])


@app.route("/api/suggest")
def api_suggest():
    """Smart-search autosuggest with typeahead instant results (blended rank)."""
    lat, lng = current_location()
    return jsonify(services.suggest(request.args.get("q", ""), lat, lng))


@app.route("/api/home")
def api_home():
    """Landing-page payload for the React frontend."""
    return jsonify({
        "popular": services.popular_products(limit=10),
        "categories": ["Clothing", "Hardware", "Electronics", "Grocery",
                       "Footwear", "Stationery", "Sports", "Home Decor"],
        "stats": {
            "shops": Shop.query.count(),
            "products": Product.query.count(),
            "orders": Order.query.count(),
        },
    })


@app.route("/api/shop/<int:shop_id>")
def api_shop(shop_id):
    s = db.session.get(Shop, shop_id) or abort(404)
    lat, lng = current_location()
    return jsonify({
        "id": s.id, "name": s.name, "category": s.category, "address": s.address,
        "contact": s.mobile, "rating": round(s.rating or 0, 1),
        "delivery": s.delivery_available, "parking": s.parking, "image": s.image,
        "lat": s.lat, "lng": s.lng, "verified": s.verified,
        "distance_km": services.haversine_km(lat, lng, s.lat, s.lng),
        "maps_url": services.maps_link(s.name, s.address),
        "products": [{
            "inventory_id": i.id, "name": i.product.name,
            "category": i.product.category, "image": i.product.image,
            "price": i.price, "stock": i.stock, "status": i.status,
        } for i in s.inventory],
    })


# ---- JSON auth API (consumed by the React SPA; shares Flask-Login session) --

def _user_json(user):
    shop = (Shop.query.filter_by(owner_id=user.id).first()
            if user.role == "owner" else None)
    return {
        "id": user.id, "name": user.name, "email": user.email,
        "phone": user.phone, "role": user.role,
        "shop_id": shop.id if shop else None,
        "shop_name": shop.name if shop else None,
    }


@app.route("/api/auth/me")
def api_me():
    if current_user.is_authenticated:
        return jsonify(_user_json(current_user))
    return jsonify(None)


@app.route("/api/auth/register", methods=["POST"])
def api_register():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email and password are required."}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "That email is already registered."}), 409
    role = data.get("role", "customer")
    if role not in ("customer", "owner"):
        role = "customer"
    user = User(name=(data.get("name") or email.split("@")[0]).strip(),
                email=email, phone=(data.get("phone") or "").strip(), role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    login_user(user)
    return jsonify(_user_json(user))


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or {}
    user = User.query.filter_by(email=(data.get("email") or "").strip().lower()).first()
    if user and user.check_password(data.get("password") or ""):
        login_user(user)
        return jsonify(_user_json(user))
    return jsonify({"error": "Invalid email or password."}), 401


@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    logout_user()
    return jsonify({"ok": True})


@app.route("/api/my/shop")
def api_my_shop():
    """Compact shopkeeper dashboard payload (owner-only)."""
    if not current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    if current_user.role != "owner":
        return jsonify({"error": "Shopkeeper account required"}), 403
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return jsonify({"shop": None})
    rev = analytics.revenue_analytics(shop.id)
    sales = analytics.sales_analytics(shop.id)
    perf = analytics.product_performance(shop.id)
    today = analytics.realtime_today(shop.id)
    return jsonify({
        "shop": {"id": shop.id, "name": shop.name, "category": shop.category,
                 "rating": shop.rating, "address": shop.address,
                 "verified": shop.verified, "delivery": shop.delivery_available},
        "revenue": rev, "sales": sales, "today": today,
        "top_selling": perf["top_selling"][:6],
        "low_stock": perf["low_stock"][:8],
        "products": [{"name": i.product.name, "price": i.price, "stock": i.stock,
                      "status": i.status, "image": i.product.image}
                     for i in shop.inventory],
    })


def _owner_shop_or_error():
    if not current_user.is_authenticated:
        return None, (jsonify({"error": "Not authenticated"}), 401)
    if current_user.role != "owner":
        return None, (jsonify({"error": "Shopkeeper account required"}), 403)
    shop = Shop.query.filter_by(owner_id=current_user.id).first()
    if not shop:
        return None, (jsonify({"error": "No shop"}), 404)
    return shop, None


@app.route("/api/order", methods=["POST"])
def api_create_order():
    """Place an order from the React storefront — either home DELIVERY or
    RESERVE & PICKUP (informs the shop in advance so they keep it ready)."""
    if not current_user.is_authenticated:
        return jsonify({"error": "Please log in to order."}), 401
    if current_user.role != "customer":
        return jsonify({"error": "Log in with a customer account to order."}), 403

    data = request.get_json(silent=True) or {}
    inv = db.session.get(Inventory, data.get("inventory_id"))
    if not inv:
        return jsonify({"error": "Product not found."}), 404

    qty = max(1, int(data.get("qty", 1)))
    if qty > inv.stock:
        return jsonify({"error": f"Only {inv.stock} in stock."}), 400

    mode = data.get("mode", "pickup")          # 'delivery' | 'pickup'
    note = (data.get("note") or "").strip()
    addr = (data.get("address") or "").strip()
    payment = data.get("payment", "cod")

    if mode == "delivery":
        order_type, where = "delivery", (addr or "Home delivery")
    else:
        order_type = "reserve"
        where = "In-store pickup" + (f" — {note}" if note else "")

    order = Order(customer_id=current_user.id, shop_id=inv.shop_id,
                  order_type=order_type, status="placed",
                  total=round(inv.price * qty, 2), address=where)
    db.session.add(order)
    db.session.flush()
    db.session.add(OrderItem(order_id=order.id, inventory_id=inv.id,
                             product_name=inv.product.name, qty=qty, price=inv.price))
    inv.stock = max(0, inv.stock - qty)        # hold the reserved/ordered units
    inv.sold = (inv.sold or 0) + qty
    db.session.add(Payment(order_id=order.id, method=payment, amount=order.total,
                           status="paid" if payment != "cod" else "pending",
                           txn_ref=f"TXN{order.id:06d}"))

    pickup_eta = None
    if order_type == "delivery":
        partner = services.assign_delivery_partner(inv.shop)
        if partner:
            order.delivery_partner_id = partner.id
            order.status = "accepted"
    db.session.commit()

    return jsonify({
        "id": order.id, "status": order.status, "type": order.order_type,
        "total": order.total, "qty": qty, "product": inv.product.name,
        "shop": inv.shop.name, "where": where,
        "message": ("Order placed for home delivery!" if order_type == "delivery"
                    else "Reserved! The shop has been notified to keep it ready for pickup."),
    })


@app.route("/api/order/<int:order_id>/action", methods=["POST"])
def api_order_action(order_id):
    """Shopkeeper order status transitions (accept/reject + Kanban moves)."""
    shop, err = _owner_shop_or_error()
    if err:
        return err
    order = db.session.get(Order, order_id)
    if not order or order.shop_id != shop.id:
        return jsonify({"error": "Order not found"}), 404
    action = (request.get_json(silent=True) or {}).get("action", "")
    transitions = {
        "accept": "accepted", "reject": "rejected", "pack": "packed",
        "dispatch": "out_for_delivery", "deliver": "delivered",
        "collect": "collected",
        # direct status set (used by Kanban drag-drop)
        "placed": "placed", "accepted": "accepted", "packed": "packed",
        "out_for_delivery": "out_for_delivery", "delivered": "delivered",
    }
    if action not in transitions:
        return jsonify({"error": "Invalid action"}), 400
    order.status = transitions[action]
    db.session.commit()
    return jsonify({"id": order.id, "status": order.status})


def _my_order_or_error(order_id):
    if not current_user.is_authenticated:
        return None, (jsonify({"error": "Not authenticated"}), 401)
    order = db.session.get(Order, order_id)
    if not order or order.customer_id != current_user.id:
        return None, (jsonify({"error": "Order not found"}), 404)
    return order, None


@app.route("/api/my/order/<int:order_id>/cancel", methods=["POST"])
def api_cancel_order(order_id):
    """Customer cancels their own order (only if not already fulfilled).
    Restores the reserved stock back to the shop."""
    order, err = _my_order_or_error(order_id)
    if err:
        return err
    if order.status in ("delivered", "collected", "cancelled", "rejected"):
        return jsonify({"error": f"This order can't be cancelled ({order.status})."}), 400
    for it in order.items:
        inv = db.session.get(Inventory, it.inventory_id) if it.inventory_id else None
        if inv:
            inv.stock = (inv.stock or 0) + it.qty
            inv.sold = max(0, (inv.sold or 0) - it.qty)
    order.status = "cancelled"
    if order.payment:
        order.payment.status = "cancelled"
    db.session.commit()
    return jsonify({"id": order.id, "status": "cancelled"})


@app.route("/api/my/order/<int:order_id>/remove", methods=["POST", "DELETE"])
def api_remove_order(order_id):
    """Remove an order from the customer's history (delete the record)."""
    order, err = _my_order_or_error(order_id)
    if err:
        return err
    # if still active, restore stock before deleting
    if order.status not in ("delivered", "collected", "cancelled", "rejected"):
        for it in order.items:
            inv = db.session.get(Inventory, it.inventory_id) if it.inventory_id else None
            if inv:
                inv.stock = (inv.stock or 0) + it.qty
                inv.sold = max(0, (inv.sold or 0) - it.qty)
    db.session.delete(order)
    db.session.commit()
    return jsonify({"ok": True, "id": order_id})


@app.route("/api/my/orders")
def api_my_orders_json():
    """The logged-in customer's own orders (with items + quantities)."""
    if not current_user.is_authenticated:
        return jsonify({"error": "Not authenticated"}), 401
    orders = sorted(current_user.orders,
                    key=lambda o: o.created_at.timestamp() if o.created_at else 0,
                    reverse=True)
    return jsonify([{
        "id": o.id, "status": o.status, "type": o.order_type,
        "total": round(o.total, 2),
        "shop": o.shop.name if o.shop else "—",
        "shop_maps_url": services.maps_link(o.shop.name, o.shop.address) if o.shop else None,
        "date": (o.created_at.strftime("%d %b %Y, %H:%M") if o.created_at else ""),
        "qty": sum(it.qty for it in o.items),
        "payment": (o.payment.method.upper() if o.payment else None),
        "items": [{"name": it.product_name, "qty": it.qty,
                   "price": round(it.price, 2)} for it in o.items],
    } for o in orders])


@app.route("/api/my/analytics")
def api_my_analytics():
    """Full business-intelligence payload for the React shopkeeper dashboard."""
    shop, err = _owner_shop_or_error()
    if err:
        return err
    data = analytics.shop_dashboard_data(shop.id)
    data["shop"] = {"id": shop.id, "name": shop.name, "category": shop.category,
                    "rating": shop.rating, "address": shop.address,
                    "contact": shop.mobile, "lat": shop.lat, "lng": shop.lng,
                    "maps_url": services.maps_link(shop.name, shop.address),
                    "verified": shop.verified, "delivery": shop.delivery_available}

    # Orders for this shop (what orders they have + quantities)
    orders = sorted(shop.orders,
                    key=lambda o: o.created_at.timestamp() if o.created_at else 0,
                    reverse=True)
    def _odist(o):
        if o.customer and o.customer.lat:
            return services.humanize_distance(
                services.haversine_km(shop.lat, shop.lng, o.customer.lat, o.customer.lng))
        return "—"

    data["orders"] = [{
        "id": o.id, "status": o.status, "type": o.order_type,
        "total": round(o.total, 2),
        "date": (o.created_at.strftime("%d %b %Y, %H:%M") if o.created_at else ""),
        "qty": sum(it.qty for it in o.items),
        "items": [{"name": it.product_name, "qty": it.qty} for it in o.items],
        "customer": (o.customer.name if o.customer else "Guest"),
        "distance": _odist(o),
        "payment": (o.payment.method.upper() if o.payment else "—"),
    } for o in orders[:40]]
    data["order_counts"] = analytics.sales_analytics(shop.id)

    # Stock status — in-stock quantities, low stock, out of stock
    in_stock, low_stock, out_stock = [], [], []
    for i in shop.inventory:
        row = {"name": i.product.name, "stock": i.stock, "price": i.price,
               "image": i.product.image}
        if i.stock <= 0:
            out_stock.append(row)
        elif i.stock <= Inventory.LOW_STOCK_THRESHOLD:
            low_stock.append(row)
        else:
            in_stock.append(row)
    data["stock"] = {
        "in_stock": sorted(in_stock, key=lambda r: -r["stock"]),
        "low_stock": sorted(low_stock, key=lambda r: r["stock"]),
        "out_of_stock": out_stock,
        "counts": {"in_stock": len(in_stock), "low_stock": len(low_stock),
                   "out_of_stock": len(out_stock),
                   "total_units": sum(i.stock for i in shop.inventory)},
    }
    return jsonify(data)


@app.route("/api/my/export.xlsx")
def api_my_export_excel():
    shop, err = _owner_shop_or_error()
    if err:
        return err
    return _build_excel(shop)


@app.route("/api/my/export.pdf")
def api_my_export_pdf():
    shop, err = _owner_shop_or_error()
    if err:
        return err
    return _build_pdf(shop)


@app.route("/api/recommendations")
@login_required
def api_recommendations():
    return jsonify(services.recommend_for_user(current_user.id))


@app.route("/api/shop/<int:shop_id>/forecast")
@role_required("owner", "admin")
def api_forecast(shop_id):
    return jsonify({
        "forecast": services.demand_forecast(shop_id),
        "restock": services.restock_suggestions(shop_id),
    })


@app.errorhandler(403)
def forbidden(e):
    return render_template("error.html", code=403,
                           message="You don't have access to this page."), 403


@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", code=404,
                           message="Page not found."), 404


# ---------------------------------------------------------------------------

def init_db():
    """Create tables and seed sample data on first run."""
    with app.app_context():
        db.create_all()
        if Shop.query.count() == 0:
            import seed
            seed.run(db)
            print("✓ Database seeded with sample stores, products & users.")


if __name__ == "__main__":
    import os
    init_db()
    port = int(os.environ.get("PORT", 5000))
    print(f"\n🛒  localConnect running at http://127.0.0.1:{port}")
    print("    Demo logins → admin@localfind.in / owner1@localfind.in / "
          "customer@localfind.in   (password: password123)\n")
    app.run(debug=True, port=port)
